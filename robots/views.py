import openpyxl
import json
from datetime import datetime, timedelta

from django.db.models import Count

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Robot

from robots.decorators import admin_required, director_required

@admin_required
@csrf_exempt
def create_robot(request):
    if request.method == 'POST':
        try:
            # Получение данных из запроса
            data = json.loads(request.body)
            model = data['model']
            version = data['version']
            created = data['created']
            
            if Robot.objects.filter(model=model, version=version, created=created).exists():
                return JsonResponse({'error': 'Robot with specified parameters already exists'}, status=400)
            
            serial = f"{model}-{version}"
            # Создание робота
            robot = Robot(serial=serial, model=model, version=version, created=created)
            robot.save()
            
            return JsonResponse({'success': 'Robot created'}, status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
    else:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
@admin_required
def list_robots(request):
   
    robot_counts = Robot.objects.values('model').annotate(model_count=Count('id'))

    robots = Robot.objects.all()
    robot_data = []

    for robot in robots:
        robot_data.append({
            'serial': robot.serial,
            'model': robot.model,
            'version': robot.version,
        })

    return JsonResponse({'robot_counts': list(robot_counts), 'robots': robot_data})

@director_required
def download_summary(request):
    # Определяем дату начала недели
    start_of_week = datetime.now() - timedelta(days=datetime.now().weekday())

    # Получаем данные о количестве роботов каждой модели за текущую неделю
    robot_summary = Robot.objects.filter(created__gte=start_of_week).values('model', 'version').annotate(total_quantity=Count('model'))
    
    # Создаем новый Excel-файл
    workbook = openpyxl.Workbook()
    
    model_sheets = {}
    
    for robot_data in robot_summary:
        model = robot_data['model']
        version = robot_data['version']
        total_quantity = robot_data['total_quantity']

        
        if model not in model_sheets:
            model_sheets[model] = workbook.create_sheet(title=model)

            # Заголовки столбцов для каждой страницы
            model_sheets[model]['A1'] = "Модель"
            model_sheets[model]['B1'] = "Версия"
            model_sheets[model]['C1'] = "Количество за неделю"

        
        row_num = model_sheets[model].max_row + 1
        model_sheets[model].cell(row=row_num, column=1, value=model)
        model_sheets[model].cell(row=row_num, column=2, value=version)
        model_sheets[model].cell(row=row_num, column=3, value=total_quantity)

    # Удаляем страницу по умолчанию
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Создаем HTTP-ответ с данными Excel-файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=info_for_dir.xlsx'

    # Сохраняем файл в HTTP-ответ
    workbook.save(response)

    return response